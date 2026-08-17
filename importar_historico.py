"""
Importa ventas históricas desde lesang_finanzas_V7.xlsx al Sheet del POS.
NO toca Mayo.
Ejecutar: python3 importar_historico.py
"""
import sys, os
sys.path.insert(0, '.')

# Token local
if not os.environ.get('GOOGLE_TOKEN_JSON'):
    token_path = os.path.join(os.path.dirname(__file__), 'token_local.json')
    if os.path.exists(token_path):
        with open(token_path, 'r') as f:
            os.environ['GOOGLE_TOKEN_JSON'] = f.read()
    else:
        print("ERROR: Crea token_local.json con el contenido de GOOGLE_TOKEN_JSON de Railway")
        sys.exit(1)

from openpyxl import load_workbook
from pos_sheets import get_creds, get_or_create_sheet
from googleapiclient.discovery import build
from datetime import datetime

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'lesang_finanzas.xlsx')
MESES_IMPORTAR = ['Febrero','Marzo','Abril','Octubre','Noviembre','Diciembre']

# Mapeo columnas Excel → POS Sheet
# Excel: Nombre(0), Talla(1), Propietario(2), Vendedor(3), Precio(4), Pago(5),
#        IVA(6), %ComB(7), $ComB(8), BaseComV(9), %ComV(10), $ComV(11),
#        Neto(12), Obs(13), Marca(14)

def limpiar(v):
    if v is None: return ''
    if isinstance(v, float) and v == int(v): return str(int(v))
    return str(v).strip()

def num(v):
    try: return float(v or 0)
    except: return 0.0

def main():
    print(f"Leyendo Excel: {EXCEL_PATH}")
    if not os.path.exists(EXCEL_PATH):
        print("ERROR: No se encontró el Excel. Asegúrate de que esté en la carpeta del proyecto.")
        sys.exit(1)

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    creds  = get_creds()
    sheets = build("sheets", "v4", credentials=creds)
    sid    = get_or_create_sheet()

    total_importadas = 0

    for mes in MESES_IMPORTAR:
        if mes not in wb.sheetnames:
            print(f"\n{mes}: no encontrado en Excel, saltando")
            continue

        ws = wb[mes]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]

        # Buscar fila de cabecera
        header_idx = None
        for i, row in enumerate(rows):
            if row[0] == 'Nombre prenda':
                header_idx = i
                break

        if header_idx is None:
            print(f"\n{mes}: sin cabecera, saltando")
            continue

        # Extraer ventas válidas
        ventas_excel = []
        for row in rows[header_idx+1:]:
            nombre = limpiar(row[0])
            if not nombre or nombre in ('TOTALES', '') or 'Gastos' in nombre:
                continue
            precio = num(row[4])
            if precio <= 0:
                continue
            ventas_excel.append(row)

        if not ventas_excel:
            print(f"\n{mes}: sin ventas, saltando")
            continue

        print(f"\n{mes}: {len(ventas_excel)} ventas a importar")

        # Verificar que el mes esté vacío en el POS Sheet
        result = sheets.spreadsheets().values().get(
            spreadsheetId=sid,
            range=f"{mes}!A3:A",
            valueRenderOption="UNFORMATTED_VALUE",
        ).execute()
        existing = result.get("values", [])
        existing_data = [r for r in existing if r and r[0] and str(r[0]).strip()]
        if existing_data:
            print(f"  ⚠ {mes} ya tiene {len(existing_data)} filas — saltando para no duplicar")
            continue

        # Construir filas para el Sheet del POS
        # Formato POS: Nombre, Talla, Propietario, Vendedor, PrecioBruto, Pago,
        #              IVA, %ComB, $ComB, BaseComV, %ComV, $ComV, Neto, Obs, Marca,
        #              Fecha, Hora, OrdenShopify, FotoLink
        pos_rows = []
        for row in ventas_excel:
            fecha_str = f"01/{mes[:3].upper()}/2025"  # fecha referencial
            pos_row = [
                limpiar(row[0]),   # Nombre prenda
                limpiar(row[1]),   # Talla
                limpiar(row[2]),   # Propietario
                limpiar(row[3]),   # Vendedor
                num(row[4]),       # Precio bruto
                limpiar(row[5]),   # Tipo pago
                num(row[6]),       # IVA
                num(row[7]),       # % com bancaria
                num(row[8]),       # $ com bancaria
                num(row[9]),       # Base com vendedor
                num(row[10]),      # % com vendedor
                num(row[11]),      # $ com vendedor
                num(row[12]),      # Neto tienda
                limpiar(row[13]),  # Observaciones
                limpiar(row[14]) if len(row)>14 else 'Tienda',  # Marca
                fecha_str,         # Fecha
                '',                # Hora
                '',                # Orden Shopify
                '',                # Foto link
            ]
            pos_rows.append(pos_row)

        # Escribir en el Sheet
        start_row = 3
        end_row = start_row + len(pos_rows) - 1
        range_str = f"{mes}!A{start_row}:S{end_row}"

        sheets.spreadsheets().values().update(
            spreadsheetId=sid,
            range=range_str,
            valueInputOption="RAW",
            body={"values": pos_rows},
        ).execute()

        print(f"  ✓ {len(pos_rows)} ventas importadas en hoja '{mes}'")
        total_importadas += len(pos_rows)

    print(f"\n{'='*50}")
    print(f"TOTAL IMPORTADO: {total_importadas} ventas")
    print(f"Meses procesados: {', '.join(MESES_IMPORTAR)}")

if __name__ == "__main__":
    main()